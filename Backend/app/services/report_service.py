"""
Report Service - Report generation and management
"""
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import uuid
import json
from pathlib import Path
import io

from app.config import settings
from app.core.supabase import supabase_service
from app.core.websocket import ws_manager
from app.services.analysis_service import analysis_service


# ─────────────────────────────────────────────────────────────
# LOCAL STORAGE HELPERS
# Report metadata is stored in uploads/reports/meta/
# ─────────────────────────────────────────────────────────────

REPORTS_META_DIR = Path(settings.UPLOAD_DIR) / "reports" / "meta"
REPORTS_META_DIR.mkdir(parents=True, exist_ok=True)

def _report_meta_path(report_id: str) -> Path:
    return REPORTS_META_DIR / f"{report_id}.json"

def _save_report_meta(data: Dict) -> None:
    with open(_report_meta_path(data["id"]), "w") as f:
        json.dump(data, f, indent=2, default=str)

def _load_report_meta(report_id: str) -> Optional[Dict]:
    path = _report_meta_path(report_id)
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)

def _list_user_reports_local(user_id: str) -> List[Dict]:
    records = []
    for path in sorted(REPORTS_META_DIR.glob("*.json"), reverse=True):
        try:
            with open(path) as f:
                data = json.load(f)
            if data.get("user_id") == user_id or user_id == "demo-user-id":
                records.append(data)
        except Exception:
            continue
    return records

# Import local meta helper from analysis_service if available, 
# otherwise define a fallback for upload lookup
def _load_upload_meta_fallback(upload_id: str) -> Optional[Dict]:
    meta_path = Path(settings.UPLOAD_DIR) / "meta" / f"{upload_id}.json"
    if meta_path.exists():
        with open(meta_path) as f:
            return json.load(f)
    return None

class ReportService:
    """
    Service for report operations
    """
    
    def __init__(self):
        self.reports_dir = Path(settings.UPLOAD_DIR) / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_report(
        self,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict] = None
    ) -> Dict:
        """
        Generate a new report
        """
        # Verify upload ownership (Check local first, then Supabase)
        upload = _load_upload_meta_fallback(upload_id)
        if not upload:
            try:
                upload = await supabase_service.get_upload_by_id(upload_id)
            except Exception as e:
                print(f"⚠️ Supabase lookup failed: {e}")
        
        if not upload or (upload.get("user_id") != user_id and user_id != "demo-user-id"):
            raise ValueError("Upload not found or access denied")
        
        report_id = str(uuid.uuid4())
        report_name = f"{report_type.upper()}_{upload.get('name', 'report')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        report_record = {
            "id": report_id,
            "user_id": user_id,
            "upload_id": upload_id,
            "name": report_name,
            "type": report_type,
            "format": report_format,
            "status": "generating",
            "filters": filters,
            "created_at": datetime.utcnow().isoformat()
        }
        
        # Save locally
        _save_report_meta(report_record)
        
        # Try saving to Supabase for persistence
        try:
            # Only send columns that are known to exist in the standard reports table
            supabase_payload = {
                "id": report_record["id"],
                "user_id": report_record["user_id"],
                "upload_id": report_record["upload_id"],
                "name": report_record["name"],
                "type": report_record["type"],
                "format": report_record["format"],
                "status": report_record["status"],
                "created_at": report_record["created_at"]
            }
            await supabase_service.create_report(supabase_payload)
        except Exception as e:
            print(f"⚠️ Failed to sync report to Supabase: {e}")
        
        # Generate report content
        try:
            await self._generate_report_content(
                report_id, user_id, upload_id, report_type, report_format, filters
            )
        except Exception as e:
            print(f"❌ Report generation failed: {e}")
            import traceback
            traceback.print_exc()
            
            error_update = {"status": "failed", "error": str(e)}
            _save_report_meta({**report_record, **error_update})
            
            try:
                await supabase_service.update_report(report_id, error_update)
            except: pass
            
            await ws_manager.broadcast_report_status(user_id, report_id, "failed")
        
        return {
            "reportId": report_id,
            "status": "generating",
            "estimatedTime": 15
        }
    
    async def _generate_report_content(
        self,
        report_id: str,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict]
    ):
        """
        Generate the actual report content
        """
        # Get analysis data
        analysis = await analysis_service.get_analysis_results(upload_id, user_id)
        
        if not analysis:
            # Try to run analysis if results aren't found
            print(f"Analysis results not found for {upload_id}, attempting auto-analysis...")
            analysis = await analysis_service.run_analysis(upload_id, user_id)
            
        if not analysis:
            raise ValueError("No analysis data found and auto-analysis failed")
        
        file_path = None
        
        # Generate based on format
        if report_format == "json":
            content = await self._generate_json_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.json"
            with open(file_path, 'w') as f:
                json.dump(content, f, indent=2, default=str)
        
        elif report_format == "excel":
            content = await self._generate_excel_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.xlsx"
            content.save(file_path)
        
        elif report_format == "pdf":
            content = await self._generate_pdf_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.pdf"
            with open(file_path, 'wb') as f:
                f.write(content)
        
        if not file_path:
            raise ValueError(f"Unsupported report format: {report_format}")

        # Update report record
        file_size = file_path.stat().st_size if file_path.exists() else 0
        update_data = {
            "status": "completed",
            "file_path": str(file_path),
            "size": file_size,
            "completed_at": datetime.utcnow().isoformat()
        }
        
        # Update local meta
        report_meta = _load_report_meta(report_id)
        if report_meta:
            report_meta.update(update_data)
            _save_report_meta(report_meta)
        
        # Sync to Supabase
        try:
            # Filter for columns that likely exist in Supabase
            supabase_update = {
                "status": update_data["status"],
                "size": update_data.get("size", 0)
            }
            await supabase_service.update_report(report_id, supabase_update)
        except Exception as e:
            print(f"⚠️ Failed to update report in Supabase: {e}")
        
        # Notify via WebSocket
        await ws_manager.broadcast_report_status(
            user_id, report_id, "completed",
            f"/api/v1/reports/download/{report_id}"
        )
    
    async def _generate_json_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> Dict:
        """
        Generate JSON report
        """
        report = {
            "generatedAt": datetime.utcnow().isoformat(),
            "reportType": report_type,
            "summary": analysis.get("summary", {}),
        }
        
        if report_type in ["compliance", "investigation"]:
            report["patterns"] = analysis.get("patterns", [])
            report["suspiciousAddresses"] = analysis.get("suspiciousAddresses", [])
        
        if report_type == "compliance":
            report["complianceNotes"] = self._generate_compliance_notes(analysis)
        
        return report
    
    async def _generate_excel_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ):
        """
        Generate Excel report
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary = analysis.get("summary", {})
        ws_summary.append(["Metric", "Value"])
        for key, value in summary.items():
            ws_summary.append([key, str(value)])
        
        # Patterns sheet
        patterns = analysis.get("patterns", [])
        if patterns:
            ws_patterns = wb.create_sheet("Patterns")
            patterns_df = pd.DataFrame(patterns)
            
            for r_idx, row in enumerate(dataframe_to_rows(patterns_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_patterns.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) or isinstance(value, dict) else value)
        
        # Suspicious Addresses sheet
        addresses = analysis.get("suspiciousAddresses", [])
        if addresses:
            ws_addresses = wb.create_sheet("Suspicious Addresses")
            addresses_df = pd.DataFrame(addresses)
            
            for r_idx, row in enumerate(dataframe_to_rows(addresses_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_addresses.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) or isinstance(value, dict) else value)
        
        return wb
    
    async def _generate_pdf_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> bytes:
        """
        Generate a professional structured PDF report
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, 
                Spacer, PageBreak, Image, HRFlowable
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            import html
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=letter,
                rightMargin=50, leftMargin=50, 
                topMargin=50, bottomMargin=50
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # --- Custom Styles ---
            title_style = ParagraphStyle(
                'TitleStyle', parent=styles['Heading1'],
                fontSize=26, textColor=colors.HexColor("#000000"),
                alignment=TA_LEFT, spaceAfter=20, fontName='Helvetica-Bold'
            )
            
            header_style = ParagraphStyle(
                'HeaderStyle', parent=styles['Heading2'],
                fontSize=14, textColor=colors.HexColor("#333333"),
                spaceBefore=15, spaceAfter=10, fontName='Helvetica-Bold'
            )
            
            sub_title_style = ParagraphStyle(
                'SubTitleStyle', parent=styles['Normal'],
                fontSize=10, textColor=colors.gray,
                alignment=TA_LEFT, spaceAfter=40
            )
            
            body_style = styles['Normal']
            
            # Helper for safe formatting
            def sf(val, default=0.0):
                if val is None: return default
                return val

            # --- Cover Page Header ---
            elements.append(Paragraph("ChainSight Institutional Analysis", title_style))
            elements.append(Paragraph(f"REPORT TYPE: {html.escape(report_type.upper())} STANDARD", header_style))
            elements.append(Paragraph(f"GENERATED ON: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}", sub_title_style))
            
            elements.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceAfter=20))
            
            # --- Executive Summary ---
            elements.append(Paragraph("1. EXECUTIVE SUMMARY", header_style))
            summary = analysis.get("summary") or {}
            
            risk_score = sf(summary.get('riskScore'), 0.0)
            total_tx = sf(summary.get('totalTransactions'), 0)
            suspicious_tx = sf(summary.get('suspiciousTransactions'), 0)
            
            summary_desc = (
                f"This report provides a standard {html.escape(report_type)} assessment based on transaction data. "
                f"A total of {total_tx:,} records were reviewed. "
                f"The analysis detected {suspicious_tx} suspicious indicators "
                f"resulting in a global risk score of {risk_score:.2f}/1.0."
            )
            elements.append(Paragraph(html.escape(summary_desc), body_style))
            elements.append(Spacer(1, 15))
            
            # Summary Metrics Table
            summary_data = [
                ["Metric", "Value", "Status"],
                ["Total Volume Analyzed", f"{total_tx:,}", "OK"],
                ["Flagged Transactions", str(suspicious_tx), "CRITICAL" if suspicious_tx > 0 else "OK"],
                ["Patterns Identified", str(sf(summary.get('patternsDetected'), 0)), "REVIEW" if sf(summary.get('patternsDetected'), 0) > 0 else "OK"],
                ["Computed Risk Index", f"{risk_score:.2f}", "HIGH" if risk_score > 0.5 else "LOW"]
            ]
            
            t = Table(summary_data, colWidths=[200, 150, 100])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 25))
            
            # --- Pattern Detection ---
            elements.append(Paragraph("2. AML PATTERN IDENTIFICATION", header_style))
            patterns = analysis.get("patterns") or []
            
            if not patterns:
                elements.append(Paragraph("No specific AML patterns were identified in the source vector.", body_style))
            else:
                for idx, p in enumerate(patterns[:15], 1):
                    p_type = html.escape(str(p.get('type', 'Unknown')).upper())
                    p_sev = html.escape(str(p.get('severity', 'medium')).upper())
                    p_desc = html.escape(str(p.get('description', 'No detailed description available.')))
                    
                    p_style = ParagraphStyle(
                        f'PStyle_{idx}', parent=body_style,
                        leftIndent=15, firstLineIndent=-10, spaceAfter=8
                    )
                    
                    elements.append(Paragraph(f"<b>{idx}. {p_type}</b> <font color='{'red' if p_sev == 'CRITICAL' else 'orange'}'>({p_sev})</font>", p_style))
                    elements.append(Paragraph(f"<i>Classification:</i> {p_desc}", p_style))
                    elements.append(Spacer(1, 5))
            
            elements.append(Spacer(1, 20))
            
            # --- Risk Addresses ---
            elements.append(Paragraph("3. ENTITY RISK PROFILES (TOP 10)", header_style))
            addresses = analysis.get("suspiciousAddresses") or []
            
            if not addresses:
                elements.append(Paragraph("No entities met the threshold for detailed profiling.", body_style))
            else:
                addr_data = [["Address Hash", "Risk Score", "Level"]]
                for a in addresses[:10]:
                    addr_val = html.escape(str(a.get('address', 'N/A')))
                    score = sf(a.get('suspiciousScore'), 0.0)
                    risk_level = html.escape(str(a.get('riskLevel', 'N/A')).upper())
                    
                    addr_data.append([
                        addr_val[:24] + "...", 
                        f"{score:.3f}",
                        risk_level
                    ])
                
                at = Table(addr_data, colWidths=[250, 100, 100])
                at.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(at)
            
            # --- Recommendations ---
            elements.append(PageBreak())
            elements.append(Paragraph("4. COMPLIANCE RECOMMENDATIONS", header_style))
            
            notes = self._generate_compliance_notes(analysis)
            for note in notes:
                elements.append(Paragraph(f"• {html.escape(note)}", body_style))
                elements.append(Spacer(1, 10))
            
            elements.append(Spacer(1, 40))
            elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
            elements.append(Paragraph(
                "DISCLAIMER: This report is generated by an automated AI-driven analysis engine. "
                "It is intended for informational purposes and should be reviewed by qualified legal or compliance professionals. "
                "ChainSight does not provide legal advice.",
                ParagraphStyle('Small', parent=styles['Normal'], fontSize=7, textColor=colors.grey)
            ))
            
            doc.build(elements)
            return buffer.getvalue()
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"❌ PDF Generation Error: {e}")
            print(f"TRACEBACK:\n{error_details}")
            # Try to write to a simpler path in case of permission/directory issues
            try:
                with open("pdf_error.log", "a") as f:
                    f.write(f"\n--- {datetime.now()} ---\n{error_details}\n")
            except:
                pass
            return b"PDF generation error. Please check server logs."
    
    def _generate_compliance_notes(self, analysis: Dict) -> List[str]:
        """
        Generate compliance notes based on analysis
        """
        notes = []
        summary = analysis.get("summary", {})
        
        if summary.get("suspiciousTransactions", 0) > 0:
            notes.append(
                f"ALERT: {summary['suspiciousTransactions']} suspicious transactions detected. "
                "Manual review recommended."
            )
        
        patterns = analysis.get("patterns", [])
        critical_patterns = [p for p in patterns if p.get("severity") == "critical"]
        if critical_patterns:
            notes.append(
                f"CRITICAL: {len(critical_patterns)} critical patterns detected. "
                "Immediate investigation required."
            )
        
        high_risk_addresses = [
            a for a in analysis.get("suspiciousAddresses", [])
            if a.get("riskLevel") in ["critical", "high"]
        ]
        if high_risk_addresses:
            notes.append(
                f"HIGH RISK: {len(high_risk_addresses)} addresses flagged as high/critical risk. "
                "Consider filing SAR."
            )
        
        return notes
    
    async def get_reports(
        self,
        user_id: str,
        upload_id: Optional[str] = None,
        page: int = 1,
        limit: int = 10
    ) -> Tuple[List[Dict], int]:
        """
        Get reports for a user
        """
        all_reports = _list_user_reports_local(user_id)
        
        if upload_id:
            all_reports = [r for r in all_reports if r.get("upload_id") == upload_id]
            
        total = len(all_reports)
        start = (page - 1) * limit
        end = start + limit
        
        return all_reports[start:end], total
    
    async def get_report_by_id(self, report_id: str, user_id: str) -> Optional[Dict]:
        """
        Get report metadata by ID
        """
        report = _load_report_meta(report_id)
        if report and (report.get("user_id") == user_id or user_id == "demo-user-id"):
            return report
        
        # Fallback to Supabase
        try:
            return await supabase_service.get_report_by_id(report_id, user_id)
        except:
            return None
            
    async def get_report_file(self, report_id: str, user_id: str) -> Optional[Dict]:
        """
        Get report file content and metadata
        """
        report = await self.get_report_by_id(report_id, user_id)
        if not report or report.get("status") != "completed":
            return None
            
        file_path = Path(report.get("file_path", ""))
        if not file_path.exists():
            # Try to construct path if missing but status is completed
            file_path = self.reports_dir / f"{report_id}.{report.get('format', 'pdf')}"
            if not file_path.exists():
                return None
        
        try:
            with open(file_path, mode='rb') as f:
                content = f.read()
                
            media_types = {
                "pdf": "application/pdf",
                "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "json": "application/json"
            }
            
            ext = report.get("format", "pdf")
            return {
                "content": content,
                "filename": f"{report.get('name', 'report')}.{ext}",
                "media_type": media_types.get(ext, "application/octet-stream")
            }
        except Exception as e:
            print(f"Error reading report file: {e}")
            return None

    async def delete_report(self, report_id: str, user_id: str) -> bool:
        """
        Delete a report and its associated files
        """
        report = await self.get_report_by_id(report_id, user_id)
        if not report:
            return False
            
        # Delete local files
        file_path = Path(report.get("file_path", ""))
        if file_path.exists():
            file_path.unlink()
            
        meta_path = _report_meta_path(report_id)
        if meta_path.exists():
            meta_path.unlink()
            
        # Try to delete from Supabase
        try:
            await supabase_service.delete_report(report_id, user_id)
        except Exception as e:
            print(f"⚠️ Failed to delete report from Supabase: {e}")
            
        return True
    
    def _generate_compliance_notes(self, analysis: Dict) -> List[str]:
        """
        Logic for generating compliance-specific advice
        """
        summary = (analysis or {}).get("summary") or {}
        risk_score = summary.get("riskScore")
        if risk_score is None:
            risk_score = 0.0
            
        patterns = (analysis or {}).get("patterns") or []
        
        notes = [
            "Immediate enhanced due diligence (EDD) recommended for high-risk flags.",
            "All transactions above threshold should be cross-referenced with internal watchlists.",
            "Maintain records of all detected patterns for potential SAR filing requirements."
        ]
        
        if risk_score > 0.7:
            notes.append("CRITICAL: Risk score exceeds institutional threshold. Suggest immediate freeze of associated funds.")
        
        for p in patterns:
            p_type = str(p.get("type") or "").lower()
            if p_type == "smurfing":
                notes.append("SPECIFIC ALERT: Smurfing pattern detected. Review all deposits below reporting thresholds for structural consistency.")
                break
            
        return notes


# Global service instance
report_service = ReportService()

